"""
デジタル寺子屋 第2回  Pythonの基礎 - 09 CSVファイルとExcelファイルを扱う

　データ処理でCSVファイルをいじったり、業務用の書類様式でExcelファイルを読み書きしたい
 というケースは多々ありますので、その処理方法をご紹介します。
 Excelが入っていない環境だとエラーになりますので、ご注意ください。

 【まず最初に！】
 1) Pycharmの最下段にある「Terminal」ボタンを押してください。
 2) コマンドプロンプト（>の横）で「pip install openpyxl」とタイプして、改行してください。
    Macの場合は「pip3 install openpyxl」と「3」を入れてください。
 3) これでExcelを扱うためのパッケージがインストールされました。importで呼び出せます。

"""

import csv                  # csvファイルの読込用パッケージのインポート
import openpyxl as excel    # Excelファイルの読込用パッケージのインポート

"""
■09-01 CSVファイルをPythonでいじる
"""
# まずはCSVファイルを作ってみる。
file = open('jikoku.csv', 'w', encoding="utf-8")
                                # jikoku.csvを書込みモード（w）で開きます。
                                # 既存のjikoku.csvは上書きで消されますので注意を。
                                # なお、モードにはr（読取専用）、a（追加書込）等があります。
                                # また文字化けのため、utf-8をエンコード方式に指定。
cf = csv.writer(file, lineterminator='\n')  # CSVファイルの書込みハンドラcfを生成
                                            # 行末の記号として\n（改行）を指定
cf.writerow([0, "特急ひのとり", "名古屋"])    # 早速書込み
cf.writerows([[1, "快速急行", "奈良"], [5, "特急", "賢島"], [6, "準急", "奈良"],
                   [10, "普通", "東生駒"]])  # どんどん書込み
file.close()    # jikoku.csvのファイルをクローズ。破損防止のため、必ず。

# 作ったCSVファイルを読みだして表示してみる。
print("[01]")
file2 = open('jikoku.csv', 'r', encoding="utf-8")
                                # jikoku.csvを今度は読込みモード（r）で開きます。
                                # また文字化けのため、utf-8をエンコード方式に指定。
data = csv.reader(file2)        # CSVファイルの読込みデータをdataに代入
for row in data:                # dataの各行を順次変数rowに取り出し
    for col in row:             # 各行の内容rowをコンマ切りで順次colに取り出し
        print(col, end='|')     # colの内容を表示
    print("*")                  # 1行分終わったら*記号で締め
file2.close()   # jikoku.csvのファイルをクローズ。破損防止のため、定番のやつ。

# 作ったCSVファイルを読みだす別の書き方。withを使ってかっちょよく。
print("[02]")
with open('jikoku.csv', 'r', encoding="utf-8") as file3:
    data = csv.reader(file3)
    for row in data:                # dataの各行を順次変数rowに取り出し
        for col in row:             # 各行の内容rowをコンマ切りで順次colに取り出し
            print(col, end='|')     # colの内容を表示
        print("*")                  # 1行分終わったら*記号で締め
# withでファイル処理の範囲が明確なのと、処理終了時にfile3をクローズしてくれるのがありがたい。

"""
■09-02 ExcelファイルをPythonでいじる
"""
# 新規にExcelファイルを作成する場合
wb = excel.Workbook()   # 新しいExcelワークブックを生成し、wbにセットする。
ws = wb.active          # wbのワークブック内でアクティブなワークシートをwsにセットする。
ws["A1"] = "まいどおおきに。"  # wsのA1セルに値を入れる。

for i in range(1, 10):  # 九九の表を作ってみる。
    for j in range(1, 10):
        p = i * j
        ws.cell(row=i+1, column=j).value = p
                        # i+1行（row）、j列（column）に積pを入れる。
                        # i+1としたのは1行目の「まいどおおきに」を避けるため。

wb.save("maido.xlsx")   # ワークブックを保存する。

# 既存のExcelファイルを読みだす場合
wb2 = excel.load_workbook("maido.xlsx")
ws2 = wb2.active

print("[03]")
for i in range(1, 10):  # さっきの九九の表を読みだしてみる。
    for j in range(1, 10):
        data = ws2.cell(row=i+1, column=j).value   #i+1行、j列のセル（cell）の値（value）を参照
        print(f"{data:2d} ",end="")    # pの値を2桁のd(digit=数値）で表示する。
    print("")   # iの段が終わったら改行（print文はデフォルトで改行する。）

wb2.close() # 保存する必要はないが、クローズは律儀にやっておく。定番、当然。